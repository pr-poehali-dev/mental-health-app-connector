"""Читает Excel-файл из S3, парсит строки и импортирует организации в БД (без дублей)"""
import json
import os
import io
import boto3
import openpyxl
import psycopg2

SCHEMA = "t_p25281756_mental_health_app_co"

def cors():
    return {"Access-Control-Allow-Origin": "*", "Access-Control-Allow-Methods": "GET, POST, OPTIONS", "Access-Control-Allow-Headers": "Content-Type"}

def get_s3():
    return boto3.client(
        "s3",
        endpoint_url="https://bucket.poehali.dev",
        aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"],
    )

def handler(event: dict, context) -> dict:
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": cors(), "body": ""}

    params = event.get("queryStringParameters") or {}

    # Режим list — показать файлы в S3
    if params.get("mode") == "list":
        s3 = get_s3()
        resp = s3.list_objects_v2(Bucket="files")
        files = [o["Key"] for o in resp.get("Contents", [])]
        return {"statusCode": 200, "headers": cors(), "body": json.dumps({"files": files})}

    # Режим inspect — показать заголовки и первые строки файла
    if params.get("mode") == "inspect":
        file_key = params.get("file")
        if not file_key:
            return {"statusCode": 400, "headers": cors(), "body": json.dumps({"error": "Укажи ?file=...&mode=inspect"})}
        s3 = get_s3()
        obj = s3.get_object(Bucket="files", Key=file_key)
        data = obj["Body"].read()
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        sheets = wb.sheetnames
        result = {}
        for sheet_name in sheets:
            ws = wb[sheet_name]
            # Ищем первую непустую строку (до 20-й)
            first_data_row = None
            for row_idx in range(1, 21):
                row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
                if any(v is not None for v in row_vals):
                    first_data_row = row_idx
                    break
            headers_raw = []
            preview_rows = []
            if first_data_row:
                headers_raw = [str(ws.cell(row=first_data_row, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
                for row_idx in range(first_data_row + 1, first_data_row + 4):
                    preview_rows.append([str(ws.cell(row=row_idx, column=c).value or "") for c in range(1, ws.max_column + 1)])
            result[sheet_name] = {
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "first_data_row": first_data_row,
                "headers": headers_raw,
                "preview": preview_rows,
            }
        return {"statusCode": 200, "headers": cors(), "body": json.dumps(result, ensure_ascii=False)}

    # Режим import — читаем конкретный файл
    file_key = params.get("file")
    if not file_key:
        return {"statusCode": 400, "headers": cors(), "body": json.dumps({"error": "Укажи ?file=path/to/file.xlsx"})}

    s3 = get_s3()
    obj = s3.get_object(Bucket="files", Key=file_key)
    data = obj["Body"].read()

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    FIELD_MAP = {
        "организация": "name", "наименование": "name", "название": "name",
        "категория": "category",
        "тип организации": "org_type",
        "для кого": "target_group", "целевая группа": "target_group",
        "краткое описание": "short_description", "описание": "short_description",
        "виды помощи": "help_types",
        "формат помощи": "help_format", "формат": "help_format",
        "условия получения": "conditions", "условия": "conditions",
        "населённый пункт": "city", "город": "city",
        "адрес": "address",
        "телефоны": "phones", "телефон": "phones",
        "email": "email",
        "сайт / соцсети": "website_social", "сайт": "website_social",
        "руководитель": "director",
        "координаты": "coordinates",
        "статус": "verification_status",
        "№": "number",
    }

    conn = psycopg2.connect(os.environ["DATABASE_URL"])
    cur = conn.cursor()
    cur.execute(f"SELECT LOWER(TRIM(name)) FROM {SCHEMA}.organizations")
    existing_names = set(r[0] for r in cur.fetchall())

    inserted = 0
    skipped = 0
    errors = []
    preview = []

    # Читаем только первый лист
    ws = wb[wb.sheetnames[0]]

    # Найти строку с заголовками (первая непустая в первых 10 строках)
    first_data_row = 1
    for row_idx in range(1, 11):
        row_vals = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_vals):
            first_data_row = row_idx
            break

    headers = [str(ws.cell(row=first_data_row, column=c).value or "").strip().lower() for c in range(1, ws.max_column + 1)]
    col_map = {}
    for i, h in enumerate(headers):
        for key, field in FIELD_MAP.items():
            if key in h:
                col_map[i] = field
                break

    for row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row + 1, values_only=True), start=first_data_row + 1):
        if not any(row):
            continue

        record = {}
        for col_idx, val in enumerate(row):
            if col_idx in col_map and val is not None:
                record[col_map[col_idx]] = str(val).strip() if val != "" else None

        name = record.get("name", "").strip()
        if not name:
            continue

        if name.lower() in existing_names:
            skipped += 1
            continue

        try:
            cur.execute(
                f"""INSERT INTO {SCHEMA}.organizations
                (number, name, category, org_type, target_group, short_description,
                 help_types, help_format, conditions, city, address, phones,
                 email, website_social, director, coordinates, verification_status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (
                    record.get("number"),
                    name,
                    record.get("category"),
                    record.get("org_type"),
                    record.get("target_group"),
                    record.get("short_description"),
                    record.get("help_types"),
                    record.get("help_format"),
                    record.get("conditions"),
                    record.get("city"),
                    record.get("address"),
                    record.get("phones"),
                    record.get("email"),
                    record.get("website_social"),
                    record.get("director"),
                    record.get("coordinates"),
                    record.get("verification_status", "pending"),
                ),
            )
            existing_names.add(name.lower())
            inserted += 1
            if len(preview) < 5:
                preview.append({"name": name, "city": record.get("city")})
        except Exception as e:
            errors.append({"row": row_idx, "name": name, "error": str(e)})

    conn.commit()
    cur.close()
    conn.close()

    return {
        "statusCode": 200,
        "headers": {**cors(), "Content-Type": "application/json"},
        "body": json.dumps({
            "ok": True,
            "inserted": inserted,
            "skipped_duplicates": skipped,
            "errors": errors,
            "preview": preview,
            "headers_found": list(set(col_map.values())),
        }, ensure_ascii=False),
    }